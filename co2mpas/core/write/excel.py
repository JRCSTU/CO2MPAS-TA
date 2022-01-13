# -*- coding: utf-8 -*-
#
# Copyright 2015-2022 European Commission (JRC);
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl
"""
Functions to write outputs on an excel file.
"""
import regex
import logging
import itertools
import schedula as sh
from collections.abc import Iterable

log = logging.getLogger(__name__)


def _clone_excel(file_name):
    from urllib.error import URLError
    import openpyxl
    try:
        from urllib.request import urlopen
        book = openpyxl.load_workbook(urlopen(file_name))
    except (ValueError, URLError):
        with open(file_name, 'rb') as file:
            book = openpyxl.load_workbook(file)
    import io
    import pandas as pd
    fd = io.BytesIO()
    # noinspection PyTypeChecker
    writer = pd.ExcelWriter(
        fd, engine='openpyxl', optimized_write=True, write_only=True
    )
    writer.book = book
    writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))
    return writer, fd


def _sort_sheets(x):
    x = x[0]
    imp = ['summary', 'dice', 'graphs', 'plan', 'nedc_h', 'nedc_l', 'wltp_h',
           'wltp_l', 'wltp_m', 'wltp_p', 'prediction', 'calibration', 'input',
           'mt', 'pa', 'ts']

    w = ()
    for i, k in enumerate(imp):
        if k in x:
            w = (i,) + _sort_sheets((x.replace(k, ''),))[0]
            break
    return w or (100,), x


def _multi_index_df2excel(writer, shname, df, index=True, **kw):
    try:
        df.to_excel(writer, shname, index=index, **kw)
    except NotImplementedError as ex:
        import pandas as pd
        if not index and isinstance(df.columns, pd.MultiIndex):
            kw = kw.copy()
            if kw.pop('header', True):
                header = pd.DataFrame([c for c in df.columns]).T
                header.to_excel(writer, shname, index=False, header=False, **kw)
                kw['startrow'] = kw.get('startrow', 0) + header.shape[0]
            values = pd.DataFrame(df.values)
            values.to_excel(writer, shname, index=False, header=False, **kw)
        else:
            raise ex


def _index_levels(index):
    # noinspection PyBroadException
    try:
        return len(index.levels)
    except Exception:
        return 1


# noinspection PyUnusedLocal
def _get_corner(df, startcol=0, startrow=0, index=True, header=True):
    if header:
        i = _index_levels(df.columns)
        startrow += i

        import pandas as pd
        if index and isinstance(df.columns, pd.MultiIndex):
            startrow += 1

    if index:
        i = _index_levels(df.index)
        startcol += i
    return startrow, startcol


def _convert_index(k):
    if not isinstance(k, Iterable):
        k = (str(k),)
    elif isinstance(k, str):
        k = (k,)
    return k


def _rangename2d(rlo, clo, rhi, chi):
    import xlrd
    return "%s:%s" % (xlrd.cellnameabs(rlo, clo), xlrd.cellnameabs(rhi, chi))


def _ranges_by_col_row(df, startrow, startcol):
    for row, i in enumerate(df.index, start=startrow):
        i = _convert_index(i)
        for col, c in enumerate(df.columns, start=startcol):
            yield i + _convert_index(c), _rangename2d(row, col, row, col)


def _ranges_by_col(df, startrow, startcol):
    for col, (k, v) in enumerate(df.items(), start=startcol):
        yield k, _rangename2d(startrow, col, startrow + len(v) - 1, col)


def _ranges_by_row(df, startrow, startcol):
    for row, (k, v) in enumerate(df.iterrows(), start=startrow):
        yield k, _rangename2d(row, startcol, row, startcol + len(v) - 1)


def _add_named_ranges(df, writer, shname, startrow, startcol, named_ranges, k0):
    ref = '!'.join([shname, '%s'])
    # noinspection PyBroadException
    try:
        define_name = writer.book.define_name

        def _create_named_range(ref_n, ref_r):
            define_name(ref % ref_n, ref % ref_r)
    except Exception:  # Use other pkg.
        define_name = writer.book.create_named_range
        scope = writer.book.index(writer.sheets[shname])

        def _create_named_range(ref_n, ref_r):
            define_name(ref_n, value=ref % ref_r, scope=scope)

    tag = ()
    if hasattr(df, 'name'):
        tag += (df.name,)

    it = ()

    if 'rows' in named_ranges and 'columns' in named_ranges:
        it += (_ranges_by_col_row(df, startrow, startcol),)
    elif 'columns' in named_ranges:
        it += (_ranges_by_col(df, startrow, startcol),)
    elif 'rows' in named_ranges:
        it += (_ranges_by_row(df, startrow, startcol),)

    for k, range_ref in itertools.chain(*it):
        k = _convert_index(k)
        if k:
            try:
                k = tag + k[k0:]
                _create_named_range(_ref_name(*k), range_ref)
            except TypeError:
                pass


def _df2excel(writer, shname, df, k0=0, named_ranges=('columns', 'rows'), **kw):
    import pandas as pd
    if isinstance(df, pd.DataFrame) and not df.empty:
        _multi_index_df2excel(writer, shname, df, **kw)

        startrow, startcol = _get_corner(df, **kw)

        if named_ranges:
            _add_named_ranges(df, writer, shname, startrow, startcol,
                              named_ranges, k0)

        return startrow, startcol


def _write_sheets(writer, sheet_name, data, down=True, **kw):
    import pandas as pd
    if isinstance(data, pd.DataFrame):
        return _df2excel(writer, sheet_name, data, **kw)
    else:
        corner = None
        for d in data:
            corner = _write_sheets(writer, sheet_name, d, down=not down, **kw)
            if corner:
                if down:
                    kw['startrow'] = d.shape[0] + corner[0] + 2
                else:
                    kw['startcol'] = d.shape[1] + corner[1] + 2
        return corner


def _sheet_name(tags):
    return '.'.join(tags)


def _ref_name(*names):
    return '_{}'.format(regex.sub(r"[\W]", "_", '.'.join(names)))


def _data_ref(ref):
    return '%s!%s' % (_sheet_name(ref[:-1]), _ref_name(ref[-1]))


def _chart2excel(writer, sheet, charts):
    import xlrd
    from openpyxl.chart import ScatterChart, Series

    sn = writer.book.sheetnames
    named_ranges = {'%s!%s' % (sn[d.localSheetId], d.name): d.value
                    for d in writer.book.defined_names.definedName}
    m, h, w = 3, 7.94, 13.55

    for i, (k, v) in enumerate(sorted(charts.items())):
        chart = ScatterChart()
        chart.height = h
        chart.width = w
        _map = {
            ('title', 'name'): ('title',),
            ('y_axis', 'name'): ('y_axis', 'title'),
            ('x_axis', 'name'): ('x_axis', 'title'),
        }
        _filter = {
            ('legend', 'position'): lambda x: x[0],
        }
        it = {s: _filter[s](o) if s in _filter else o
              for s, o in sh.stack_nested_keys(v['set'])}

        for s, o in sh.map_dict(_map, it).items():
            c = chart
            for j in s[:-1]:
                c = getattr(c, j)
            setattr(c, s[-1], o)

        for s in v['series']:
            xvalues = named_ranges[_data_ref(s['x'])]
            values = named_ranges[_data_ref(s['y'])]
            series = Series(values, xvalues, title=s['label'])
            chart.series.append(series)

        n = int(i / m)
        j = i - n * m

        sheet.add_chart(chart, xlrd.cellname(15 * j, 8 * n))


def write_to_excel(dfs, output_template):
    """
    Writes DataFrames to excel.

    :param dfs:
        DataFrames of vehicle output report.
    :type dfs: dict[str, pandas.DataFrame]

    :param output_template:
        Template output.
    :type output_template: str

    :return:
        Excel output file.
    :rtype: io.BytesIO
    """
    log.debug(
        'Writing into xl-file based on template(%s)...', output_template
    )
    writer, fd = _clone_excel(output_template)

    calculate_sheets, charts = sorted(writer.sheets), []
    for k, v in sorted(dfs.items(), key=_sort_sheets):
        if not k.startswith('graphs.'):
            down = True
            if k.endswith('pa'):
                kw = {'named_ranges': ('rows',), 'index': True, 'k0': 1}
            elif k.endswith('ts'):
                kw = {'named_ranges': ('columns',), 'index': False, 'k0': 1}
            elif k.endswith('proc_info'):
                down = False
                kw = {'named_ranges': ()}
            else:
                kw = {}
            _write_sheets(writer, k, v, down=down, **kw)
        else:
            try:
                sheet = writer.book.add_worksheet(k)
            except AttributeError:
                sheet = writer.book.create_sheet(title=k)
            charts.append((sheet, v))

    for sheet, v in charts:
        _chart2excel(writer, sheet, v)

    writer.save()
    return fd
